import openpyxl
import os
from kivy.app import App
from kivy.core.window import Window
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.checkbox import CheckBox
from kivy.uix.dropdown import DropDown
from kivy.uix.label import Label
from kivy.uix.tabbedpanel import TabbedPanel
from kivy.uix.textinput import TextInput
from kivy.uix.widget import Widget


class Container(TabbedPanel):
    def start(self):
        filepath = 'bore.xlsx'
        if not os.path.exists(filepath):  # создаётся эксель-файл, если не существует
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # запись категорий в столбцы; для записи в строки - закомментировать
            sheet.append(['Номер объекта', 'Номер скважины', 'Дата проходки', 'Абс. отметка', 'Глубина скважины', 'Глубина кровли ИГЭ', 
                          'Описание ИГЭ',  'Вид образца', 'Глубина образца', 'Глубина подошвы'])
            # конец участка для закомментирования при записи в строках
            
            workbook.save(filepath)
        else:
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            # запись категорий в столбцы; для записи в строки - закомментировать
            sheet.append(['Номер объекта', 'Номер скважины', 'Дата проходки', 'Абс. отметка', 'Глубина скважины', 'Глубина кровли ИГЭ', 
                        'Описание слоя',  'Вид образца', 'Глубина образца', 'Глубина подошвы'])
            # конец участка для закомментирования при записи в строках

            workbook.save(filepath)

    def submit(self):
        self.ids.lbl_obj_nm.text = 'Номер объекта: '+self.ids.obj_nm.text
        self.ids.lbl_bore_nm.text = 'Номер скважины: '+self.ids.bore_nm.text
        self.ids.lbl_bore_date.text = 'Дата проходки: '+self.ids.bore_date.text
        self.ids.lbl_bore_abs.text = 'Абс. отметка: '+self.ids.bore_abs.text
        self.ids.lbl_bore_depth.text = 'Глубина скважины: '+self.ids.bore_depth.text
        self.ids.lbl_layer_roof.text = 'Глубина кровли ИГЭ: '+self.ids.layer_roof.text
        self.ids.lbl_layer_descr.text = 'Описание ИГЭ: '+self.ids.layer_descr.text
        self.ids.lbl_sample_type.text = 'Вид образца: '+self.ids.sample_type.text
        self.ids.lbl_sample_depth.text = 'Глубина образца: '+self.ids.sample_depth.text
        self.ids.lbl_layer_depth.text = 'Глубина подошвы ИГЭ: '+self.ids.layer_depth.text        

    def write(self):
        filepath = 'bore.xlsx'
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        # запись категорий в столбцы
        sheet.append([self.ids.obj_nm.text, self.ids.bore_nm.text, self.ids.bore_date.text, self.ids.bore_abs.text, 
                      self.ids.bore_depth.text, self.ids.layer_roof.text, self.ids.layer_descr.text, self.ids.sample_type.text, 
                      self.ids.sample_depth.text, self.ids.layer_depth.text])

        # запись категорий в строки
        # sheet.append(['Номер объекта', self.ids.obj_nm.text])
        # sheet.append(['Номер скважины', self.ids.bore_nm.text])
        # sheet.append(['Дата проходки', self.ids.bore_date.text])
        # sheet.append(['Абс. отметка', self.ids.bore_abs.text])
        # sheet.append(['Глубина скважины', self.ids.bore_depth.text])
        # sheet.append(['Глубина кровли ИГЭ', self.ids.layer_roof.text])
        # sheet.append(['Описание ИГЭ', self.ids.layer_descr.text])
        # sheet.append(['Вид образца', self.ids.sample_type.text])
        # sheet.append(['Глубина образца', self.ids.sample_depth.text])
        # sheet.append(['Глубина подошвы', self.ids.layer_depth.text])

        workbook.save(filepath)

class BoreLogApp(App):
    def build(self):
        return Container()
        

if __name__ == '__main__':
    BoreLogApp().run()